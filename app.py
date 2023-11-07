'''
Os imports principais.
'''
import datetime as dt
import sys
from os import getcwd, chdir
from time import sleep
from os.path import join, abspath
from openpyxl import load_workbook
from winotify import Notification


atual_dir = getcwd()
sys.path.append(atual_dir)
if getattr(sys, "frozen", False):
    chdir(sys._MEIPASS)

def ver_eventos():
    arquivo_excel = abspath(join(atual_dir, 'Calendário de Datas Comemorativas - 2023.xlsx'))
    icon = abspath(join(atual_dir, 'icon.png'))

    wb = load_workbook(arquivo_excel)
    sheet = wb.active

    agora = (dt.datetime.now()).date()
    eventos_proximos = []

    for row in sheet.iter_rows(min_row=2):
        data = (row[0].value).date()
        evento = row[1].value
        tipo = row[2].value


        if agora <= data <= agora + dt.timedelta(days=7) or data == agora:
            eventos_proximos.append((evento, data.strftime("%d/%m/%Y"), tipo))


    if eventos_proximos:
        EVENTO = ', '.join(f'{evento} em {data} do tipo {tipo}' for evento, data, tipo in eventos_proximos)
        notif = Notification(app_id='Notificador de Eventos',
                            title='Datas Comemorativas, Clique em mim para ver a planilha',
                            msg=f'Olá, os eventos mais próximos são: {EVENTO}',
                            duration='long',
                            icon=icon,
                            launch=arquivo_excel
                            )
        notif.show()

while True:
    hora_agora = dt.datetime.now().time()
    if (hora_agora.hour == 11 and hora_agora.minute == 00) or (hora_agora.hour == 16 and hora_agora.minute == 00):
        ver_eventos()
    sleep(60)
